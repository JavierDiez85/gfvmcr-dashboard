#!/usr/bin/env python3
"""
Genera plantilla Excel para configuración TPV con soporte de cambios históricos de comisiones.
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Load data ──
with open(os.path.expanduser('~/Downloads/tpv_export.json'), 'r') as f:
    data = json.load(f)

clients = data['clients']
agents = data['agents'] or []

wb = Workbook()

# ── Colors ──
BLUE = '0073EA'
GREEN = '00B875'
RED = 'E53935'
ORANGE = 'FF7043'
PURPLE = '9B51E0'
GRAY = 'F5F6FA'
WHITE = 'FFFFFF'
DARK = '1A1C2E'

hdr_font = Font(name='Calibri', bold=True, size=10, color=WHITE)
hdr_fill_blue = PatternFill(start_color=BLUE, end_color=BLUE, fill_type='solid')
hdr_fill_green = PatternFill(start_color=GREEN, end_color=GREEN, fill_type='solid')
hdr_fill_orange = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type='solid')
hdr_fill_purple = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type='solid')
hdr_fill_red = PatternFill(start_color=RED, end_color=RED, fill_type='solid')
hdr_fill_dark = PatternFill(start_color=DARK, end_color=DARK, fill_type='solid')
data_fill = PatternFill(start_color=GRAY, end_color=GRAY, fill_type='solid')
warn_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='D0D4E4'),
    right=Side(style='thin', color='D0D4E4'),
    top=Side(style='thin', color='D0D4E4'),
    bottom=Side(style='thin', color='D0D4E4')
)

def style_header(ws, row, cols, fill=hdr_fill_blue):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_data_row(ws, row, cols, highlight=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name='Calibri', size=9)
        cell.border = thin_border
        if highlight:
            cell.fill = warn_fill
        cell.alignment = Alignment(horizontal='center' if c > 2 else 'left', vertical='center')


# ═══════════════════════════════════════
# SHEET 1: Aux_Data (Client Rates)
# ═══════════════════════════════════════
ws1 = wb.active
ws1.title = 'Aux_Data'

# Headers
headers_aux = [
    'Cliente_Norm', 'Cliente_Display', 'Trans_Name', 'Promotor', 'Entidad',
    'Active_Desde', 'Active_Hasta', 'Factor_IVA',
    'Efevoo_TC', 'Efevoo_TD', 'Efevoo_Amex', 'Efevoo_TI',
    'Salem_TC', 'Salem_TD', 'Salem_Amex', 'Salem_TI',
    'Convenia_TC', 'Convenia_TD', 'Convenia_Amex', 'Convenia_TI',
    'Comisionista_TC', 'Comisionista_TD', 'Comisionista_Amex', 'Comisionista_TI'
]

# Title row
ws1.merge_cells('A1:X1')
ws1['A1'] = 'CONFIGURACION DE CLIENTES — Tasas de comision actuales (CONTADO)'
ws1['A1'].font = Font(name='Calibri', bold=True, size=12, color=BLUE)
ws1['A1'].alignment = Alignment(horizontal='center')

# Group headers row 2
group_headers = [
    (1, 5, 'DATOS DEL CLIENTE', hdr_fill_dark),
    (6, 8, 'CONTRATO', hdr_fill_purple),
    (9, 12, 'EFEVOO', hdr_fill_blue),
    (13, 16, 'SALEM', hdr_fill_green),
    (17, 20, 'CONVENIA', hdr_fill_orange),
    (21, 24, 'COMISIONISTA', hdr_fill_red),
]
for start, end, label, fill in group_headers:
    ws1.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
    cell = ws1.cell(row=2, column=start, value=label)
    cell.font = hdr_font
    cell.fill = fill
    cell.alignment = Alignment(horizontal='center')

# Column headers row 3
for i, h in enumerate(headers_aux, 1):
    cell = ws1.cell(row=3, column=i, value=h)
    cell.font = Font(name='Calibri', bold=True, size=9, color=DARK)
    cell.fill = PatternFill(start_color='E8ECFA', end_color='E8ECFA', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

# Key mapping from compact JSON
key_map = {
    'Cliente_Norm': 'n', 'Cliente_Display': 'd', 'Promotor': 'p', 'Factor_IVA': 'fi',
    'Efevoo_TC': 'etc', 'Efevoo_TD': 'etd', 'Efevoo_Amex': 'ea', 'Efevoo_TI': 'eti',
    'Salem_TC': 'stc', 'Salem_TD': 'std', 'Salem_Amex': 'sa', 'Salem_TI': 'sti',
    'Convenia_TC': 'ctc', 'Convenia_TD': 'ctd', 'Convenia_Amex': 'ca', 'Convenia_TI': 'cti',
    'Comisionista_TC': 'ktc', 'Comisionista_TD': 'ktd', 'Comisionista_Amex': 'ka', 'Comisionista_TI': 'kti'
}

# Data rows
for idx, c in enumerate(clients):
    row = idx + 4
    has_any_rate = any(float(c.get(key_map.get(h, ''), 0) or 0) > 0 for h in headers_aux[8:])

    for col, h in enumerate(headers_aux, 1):
        k = key_map.get(h)
        val = c.get(k, '') if k else ''
        if h == 'Trans_Name':
            val = c.get('d', '')
        elif h == 'Entidad':
            val = ''
        elif h == 'Active_Desde' or h == 'Active_Hasta':
            val = ''

        cell = ws1.cell(row=row, column=col, value=val if val != 0 else 0)
        cell.font = Font(name='Calibri', size=9)
        cell.border = thin_border

        if not has_any_rate:
            cell.fill = warn_fill  # highlight clients without rates

        if col >= 9:  # rate columns
            cell.number_format = '0.0000%' if isinstance(val, (int, float)) and val > 0 else '@'
            cell.alignment = Alignment(horizontal='center')

# Add empty rows for new clients
for i in range(20):
    row = len(clients) + 4 + i
    for col in range(1, len(headers_aux) + 1):
        cell = ws1.cell(row=row, column=col)
        cell.fill = PatternFill(start_color='E8FAF2', end_color='E8FAF2', fill_type='solid')
        cell.border = thin_border
        if col == 8:
            cell.value = 1.16  # default factor_iva

# Column widths
widths = [22, 22, 22, 14, 10, 12, 12, 10] + [12] * 16
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.freeze_panes = 'A4'


# ═══════════════════════════════════════
# SHEET 2: Cambios_Comisiones (NUEVO)
# ═══════════════════════════════════════
ws2 = wb.create_sheet('Cambios_Comisiones')

# Title
ws2.merge_cells('A1:H1')
ws2['A1'] = 'HISTORICO DE CAMBIOS DE COMISION — Llena esta hoja si algun cliente cambio de comision durante 2025'
ws2['A1'].font = Font(name='Calibri', bold=True, size=11, color=RED)
ws2['A1'].alignment = Alignment(horizontal='center')

ws2.merge_cells('A2:H2')
ws2['A2'] = 'El sistema usara la Tasa_Anterior para transacciones ANTES de Fecha_Cambio, y la Tasa_Nueva para transacciones DESPUES'
ws2['A2'].font = Font(name='Calibri', size=9, color='666666', italic=True)
ws2['A2'].alignment = Alignment(horizontal='center')

headers_cambios = [
    'Cliente', 'Fecha_Cambio', 'Campo',
    'Tasa_Anterior', 'Tasa_Nueva',
    'Notas'
]

# Headers row 3
for i, h in enumerate(headers_cambios, 1):
    cell = ws2.cell(row=3, column=i, value=h)
style_header(ws2, 3, len(headers_cambios), hdr_fill_red)

# Instructions row 4
examples = [
    'EJEMPLO CAFE',
    '2025-06-15',
    'Efevoo_TC',
    0.018,
    0.025,
    'Renegociacion de contrato'
]
for i, val in enumerate(examples, 1):
    cell = ws2.cell(row=4, column=i, value=val)
    cell.font = Font(name='Calibri', size=9, italic=True, color='999999')
    cell.border = thin_border

# Valid field names reference
ws2.cell(row=6, column=1, value='Campos validos para columna "Campo":').font = Font(bold=True, size=9)
valid_fields = [
    'Efevoo_TC', 'Efevoo_TD', 'Efevoo_Amex', 'Efevoo_TI',
    'Salem_TC', 'Salem_TD', 'Salem_Amex', 'Salem_TI',
    'Convenia_TC', 'Convenia_TD', 'Convenia_Amex', 'Convenia_TI',
    'Comisionista_TC', 'Comisionista_TD', 'Comisionista_Amex', 'Comisionista_TI',
    'Factor_IVA'
]
for i, f in enumerate(valid_fields):
    r = 7 + i
    ws2.cell(row=r, column=1, value=f).font = Font(name='Calibri', size=8, color='666666')

# Empty rows for data (starting row 5, skip example at 4)
for row in range(5, 55):
    for col in range(1, len(headers_cambios) + 1):
        cell = ws2.cell(row=row, column=col)
        cell.border = thin_border

# Column widths
for i, w in enumerate([25, 15, 20, 14, 14, 30], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = 'A4'


# ═══════════════════════════════════════
# SHEET 3: Agentes
# ═══════════════════════════════════════
ws3 = wb.create_sheet('Agentes')

ws3.merge_cells('A1:C1')
ws3['A1'] = 'AGENTES — Equipo comercial'
ws3['A1'].font = Font(name='Calibri', bold=True, size=11, color=PURPLE)

headers_agentes = ['Agentes', 'Siglas', 'Comision por ventas']
for i, h in enumerate(headers_agentes, 1):
    ws3.cell(row=2, column=i, value=h)
style_header(ws3, 2, 3, hdr_fill_purple)

for idx, ag in enumerate(agents):
    row = idx + 3
    ws3.cell(row=row, column=1, value=ag.get('nombre', '')).border = thin_border
    ws3.cell(row=row, column=2, value=ag.get('siglas', '')).border = thin_border
    ws3.cell(row=row, column=3, value=ag.get('comision_ventas', 0)).border = thin_border

# Extra empty rows
for row in range(len(agents) + 3, len(agents) + 8):
    for col in range(1, 4):
        ws3.cell(row=row, column=col).border = thin_border
        ws3.cell(row=row, column=col).fill = PatternFill(start_color='F3EAFD', end_color='F3EAFD', fill_type='solid')

for i, w in enumerate([25, 10, 20], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════
# SHEET 4: Instrucciones
# ═══════════════════════════════════════
ws4 = wb.create_sheet('Instrucciones')

instructions = [
    ('PLANTILLA DE CONFIGURACION TPV — INSTRUCCIONES', Font(name='Calibri', bold=True, size=14, color=BLUE)),
    ('', None),
    ('HOJA 1: Aux_Data', Font(name='Calibri', bold=True, size=12, color=DARK)),
    ('Contiene TODOS los clientes actuales con sus tasas de comision vigentes.', Font(name='Calibri', size=10)),
    ('• Los clientes resaltados en AMARILLO no tienen comisiones configuradas — llenar sus tasas', Font(name='Calibri', size=10, color=ORANGE)),
    ('• Las filas verdes al final son para NUEVOS CLIENTES — agregar nombre y tasas', Font(name='Calibri', size=10, color=GREEN)),
    ('• Las tasas se expresan en decimal: 0.025 = 2.5%, 0.018 = 1.8%', Font(name='Calibri', size=10)),
    ('• Factor_IVA por defecto es 1.16 (16% IVA)', Font(name='Calibri', size=10)),
    ('', None),
    ('HOJA 2: Cambios_Comisiones  ★ NUEVA ★', Font(name='Calibri', bold=True, size=12, color=RED)),
    ('Esta hoja es para clientes que CAMBIARON de comision durante 2025.', Font(name='Calibri', size=10)),
    ('', None),
    ('COMO LLENARLA:', Font(name='Calibri', bold=True, size=10)),
    ('1. Cliente: Nombre exacto del cliente (como aparece en Aux_Data)', Font(name='Calibri', size=10)),
    ('2. Fecha_Cambio: Fecha en que la comision cambio (YYYY-MM-DD, ej: 2025-06-15)', Font(name='Calibri', size=10)),
    ('3. Campo: Cual tasa cambio (ej: Efevoo_TC, Salem_TD, etc.)', Font(name='Calibri', size=10)),
    ('4. Tasa_Anterior: La tasa que tenia ANTES del cambio (decimal)', Font(name='Calibri', size=10)),
    ('5. Tasa_Nueva: La tasa que tiene DESPUES del cambio (decimal)', Font(name='Calibri', size=10)),
    ('6. Notas: Comentario opcional', Font(name='Calibri', size=10)),
    ('', None),
    ('EJEMPLO:', Font(name='Calibri', bold=True, size=10, color=RED)),
    ('Si "CAFE CENTRAL" tenia Efevoo_TC = 0.018 hasta junio 2025 y luego subio a 0.025:', Font(name='Calibri', size=10)),
    ('→ Cliente: CAFE CENTRAL | Fecha_Cambio: 2025-07-01 | Campo: Efevoo_TC | Anterior: 0.018 | Nueva: 0.025', Font(name='Calibri', size=10, color=RED)),
    ('', None),
    ('Si cambio MAS DE UN campo, poner una fila por cada campo que cambio.', Font(name='Calibri', size=10, bold=True)),
    ('En Aux_Data dejar la tasa ACTUAL (la nueva). El historico se lee de esta hoja.', Font(name='Calibri', size=10, bold=True)),
    ('', None),
    ('HOJA 3: Agentes', Font(name='Calibri', bold=True, size=12, color=PURPLE)),
    ('Lista de agentes comerciales. Agregar nuevos o modificar existentes.', Font(name='Calibri', size=10)),
    ('', None),
    ('COMO SUBIR:', Font(name='Calibri', bold=True, size=12, color=GREEN)),
    ('1. Llenar todas las hojas necesarias', Font(name='Calibri', size=10)),
    ('2. Ir al dashboard → Configuracion → Subir Config TPV', Font(name='Calibri', size=10)),
    ('3. Seleccionar este archivo Excel', Font(name='Calibri', size=10)),
    ('4. El sistema procesara Aux_Data, Cambios_Comisiones y Agentes automaticamente', Font(name='Calibri', size=10)),
]

for i, (text, font) in enumerate(instructions, 1):
    cell = ws4.cell(row=i, column=1, value=text)
    if font:
        cell.font = font

ws4.column_dimensions['A'].width = 100


# ═══════════════════════════════════════
# SAVE
# ═══════════════════════════════════════
output_path = os.path.expanduser('~/Downloads/TPV_Config_Template_2025.xlsx')
wb.save(output_path)
print(f'Template saved to: {output_path}')
print(f'Clients: {len(clients)}, Agents: {len(agents)}')
print(f'Clients without rates: {sum(1 for c in clients if all(float(c.get(k,0) or 0)==0 for k in ["etc","stc","ctc","ktc"]))}')
